# -*- coding: utf-8 -*-
"""
일일 보고서 PDF 생성기
Excel 파일에서 데이터를 읽어 PDF 보고서를 자동 생성
"""

import os
import html
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# 한글 폰트 등록
def register_korean_font():
    """한글 폰트(맑은 고딕) 등록"""
    font_paths = [
        "C:/Windows/Fonts/malgun.ttf",  # 맑은 고딕
        "C:/Windows/Fonts/NanumGothic.ttf",  # 나눔고딕
        "C:/Windows/Fonts/gulim.ttc",  # 굴림
    ]
    
    for font_path in font_paths:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont('Korean', font_path))
                return 'Korean'
            except Exception:
                continue
    
    # 폰트를 찾지 못한 경우 기본 폰트 사용
    return 'Helvetica'


class DailyReportData:
    """일일 보고서 데이터 클래스"""
    
    def __init__(self):
        self.report_date = datetime.now().strftime('%Y-%m-%d')
        self.daily_count = 0          # 하루 통관건수
        self.daily_fee = 0            # 하루 수수료
        self.monthly_count = 0        # 월 누계 통관건수
        self.monthly_fee = 0          # 월 누계 수수료
        self.receivables = []         # 미수금 리스트 [{'company': str, 'amount': int}, ...]
        self.advances = []            # 대납금 리스트 [{'company': str, 'amount': int}, ...]
        self.issues = ""              # 통관, 인사, 영업 특이사항
    
    def add_receivable(self, company: str, amount: int, expected_date: str = "", is_confirmed: bool = False):
        """미수금 항목 추가"""
        self.receivables.append({
            'company': company, 
            'amount': int(amount or 0),
            'expected_date': expected_date,
            'is_confirmed': is_confirmed
        })
    
    def add_advance(self, company: str, bl_no: str, amount: int, expected_date: str = "", is_confirmed: bool = False):
        """대납금 항목 추가"""
        self.advances.append({
            'company': company, 
            'bl_no': bl_no,
            'amount': int(amount or 0),
            'expected_date': expected_date,
            'is_confirmed': is_confirmed
        })
    
    def get_total_receivables(self) -> int:
        """미수금 합계"""
        return sum(int(item.get('amount', 0)) for item in self.receivables)
    
    def get_total_advances(self) -> int:
        """대납금 합계"""
        return sum(int(item.get('amount', 0)) for item in self.advances)


class DailyReportGenerator:
    """일일 보고서 PDF 생성기"""
    
    # 색상 팔레트 (전문적인 보고서 스타일) - 메일과 동일한 네이비 통일
    PRIMARY_COLOR = colors.HexColor('#1f3a58')      # 진한 네이비
    SECONDARY_COLOR = colors.HexColor('#2b6cb0')    # 블루
    HEADER_BG = colors.HexColor('#e2e8f0')          # 연한 그레이
    ALT_ROW_BG = colors.HexColor('#f7fafc')         # 매우 연한 그레이
    
    def __init__(self):
        self.font_name = register_korean_font()
        self._setup_styles()
    
    def _setup_styles(self):
        """스타일 설정"""
        self.styles = getSampleStyleSheet()
        
        # 제목 스타일
        self.styles.add(ParagraphStyle(
            name='ReportTitle',
            fontName=self.font_name,
            fontSize=22,
            textColor=self.PRIMARY_COLOR,
            spaceAfter=8,
            alignment=1,  # 중앙 정렬
            leading=26
        ))
        
        # 부제목 스타일 (보고일자 - 오른쪽 정렬)
        self.styles.add(ParagraphStyle(
            name='ReportSubtitle',
            fontName=self.font_name,
            fontSize=11,
            textColor=colors.gray,
            spaceAfter=10,
            alignment=2  # 오른쪽 정렬
        ))
        
        # 섹션 헤더 스타일
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            fontName=self.font_name,
            fontSize=12,
            textColor=self.PRIMARY_COLOR,
            spaceBefore=8,
            spaceAfter=4,
            borderPadding=3
        ))
        
        # 일반 텍스트 스타일
        self.styles.add(ParagraphStyle(
            name='ReportBody',
            fontName=self.font_name,
            fontSize=10,
            textColor=colors.black,
            spaceAfter=8
        ))
        
        # 테이블 헤더 스타일
        self.styles.add(ParagraphStyle(
            name='TableHeader',
            fontName=self.font_name,
            fontSize=10,
            textColor=colors.white
        ))
        
        # 테이블 셀 스타일
        self.styles.add(ParagraphStyle(
            name='TableCell',
            fontName=self.font_name,
            fontSize=10,
            textColor=colors.black
        ))
    
    def _format_currency(self, amount: int) -> str:
        """금액 포맷 (천단위 콤마)"""
        return f"{amount:,}원"
    
    def _create_summary_table(self, data: DailyReportData) -> Table:
        """통관 요약 테이블 생성"""
        table_data = [
            ['구분', '건수', '수수료(공급가)'],
            ['당일 실적', f"{data.daily_count}건", self._format_currency(data.daily_fee)],
            ['월 누계', f"{data.monthly_count}건", self._format_currency(data.monthly_fee)],
        ]
        
        table = Table(table_data, colWidths=[60*mm, 50*mm, 60*mm])
        table.setStyle(TableStyle([
            # 헤더 스타일
            ('BACKGROUND', (0, 0), (-1, 0), self.PRIMARY_COLOR),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), self.font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            # 테두리
            ('GRID', (0, 0), (-1, -1), 1, self.HEADER_BG),
            ('BOX', (0, 0), (-1, -1), 2, self.PRIMARY_COLOR),
            # 행 배경색 번갈아
            ('BACKGROUND', (0, 1), (-1, 1), colors.white),
            ('BACKGROUND', (0, 2), (-1, 2), self.ALT_ROW_BG),
        ]))
        
        return table
    
    def _create_list_table(self, title: str, items: list, total: int, is_warning: bool = False) -> list:
        """리스트 테이블 생성 (미수금/대납금용)"""
        elements = []
        
        # 섹션 헤더
        header_color = self.PRIMARY_COLOR
        elements.append(Paragraph(f"■ {title}", self.styles['SectionHeader']))
        
        if not items:
            elements.append(Paragraph("해당 항목 없음", self.styles['ReportBody']))
            return elements
        
        # 테이블 헤더 및 데이터 구성
        # BL 번호 존재 여부 확인 (첫 번째 데이터의 키로 확인)
        has_bl = any('bl_no' in item for item in items)
        
        status_label = "확인"
        
        if has_bl:
            # BL 번호 포함 (6열)
            # No(10), Company(45), BL(35), Amount(30), Date(30), Status(20) = 170
            col_widths = [10*mm, 45*mm, 35*mm, 30*mm, 30*mm, 20*mm]
            table_data = [['No', '업체명', 'BL 번호', '금액', '입금예정일', status_label]]
        else:
            # 기존 포맷 (5열)
            # No(10), Company(60), Amount(40), Date(40), Status(20) = 170
            col_widths = [10*mm, 60*mm, 40*mm, 40*mm, 20*mm]
            table_data = [['No', '업체명', '금액', '입금예정일', status_label]]
        
        for idx, item in enumerate(items, 1):
            status_text = "■" if item.get('is_confirmed', False) else "□"
            row = [str(idx), item['company']]
            
            if has_bl:
                row.append(item.get('bl_no', ''))
            
            row.extend([
                self._format_currency(item['amount']),
                item.get('expected_date', ''),
                status_text
            ])
            table_data.append(row)
        
        # 합계 행
        if has_bl:
            # BL 번호가 있으면 합계 행도 맞춰야 함 (빈 셀 하나 더 추가)
            table_data.append(['', '합계', '', self._format_currency(total), '', ''])
            currency_col_idx = 3 # 금액 열 인덱스 (0-base)
        else:
            table_data.append(['', '합계', self._format_currency(total), '', ''])
            currency_col_idx = 2

        table = Table(table_data, colWidths=col_widths)
        
        style_commands = [
            # 헤더 스타일
            ('BACKGROUND', (0, 0), (-1, 0), header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), self.font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (currency_col_idx, 0), (currency_col_idx, -1), 'RIGHT'), # 금액 우측 정렬
            ('ALIGN', (-2, 0), (-1, -1), 'CENTER'),  # 입금예정일, 확인 중앙 정렬
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # 헤더 전체 중앙 정렬 강제
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            # 테두리
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('BOX', (0, 0), (-1, -1), 1.5, header_color),
            # 합계 행 스타일
            ('BACKGROUND', (0, -1), (-1, -1), self.HEADER_BG),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('TEXTCOLOR', (1, -1), (1, -1), self.PRIMARY_COLOR),
        ]
        
        # 행 배경색 번갈아
        for i in range(1, len(table_data) - 1):
            if i % 2 == 0:
                style_commands.append(('BACKGROUND', (0, i), (-1, i), self.ALT_ROW_BG))
        
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        
        return elements
    
    def generate_pdf(self, data: DailyReportData, output_path: str) -> str:
        """PDF 보고서 생성"""
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )
        
        elements = []
        
        # 제목
        elements.append(Paragraph("일일 업무 보고서", self.styles['ReportTitle']))
        
        # 날짜
        elements.append(Paragraph(
            f"보고일자: {data.report_date}",
            self.styles['ReportSubtitle']
        ))
        
        # 구분선
        elements.append(HRFlowable(
            width="100%",
            thickness=2,
            color=self.PRIMARY_COLOR,
            spaceAfter=8
        ))
        
        # 통관 실적 섹션
        elements.append(Paragraph("■ 통관 실적", self.styles['SectionHeader']))
        elements.append(self._create_summary_table(data))
        elements.append(Spacer(1, 5*mm))
        
        # 미수금 섹션
        elements.extend(self._create_list_table(
            "미수금 현황",
            data.receivables,
            data.get_total_receivables(),
            is_warning=True
        ))
        elements.append(Spacer(1, 3*mm))
        
        # 대납금 섹션
        elements.extend(self._create_list_table(
            "대납금 현황",
            data.advances,
            data.get_total_advances(),
            is_warning=False
        ))
        
        if data.issues:
            elements.append(Spacer(1, 3*mm))
            # 특이사항 제목 (내용 박스와 겹침 방지)
            elements.append(Paragraph(
                "■ 통관, 인사, 영업 특이사항",
                ParagraphStyle(
                    name='IssuesHeader',
                    fontName=self.font_name,
                    fontSize=12,
                    textColor=self.PRIMARY_COLOR,
                    spaceBefore=8,
                    spaceAfter=20,
                    wordWrap='CJK'
                )
            ))
            elements.append(Paragraph(
                html.escape(data.issues).replace('\n', '<br/>'),
                ParagraphStyle(
                    name='Issues',
                    parent=self.styles['Normal'],
                    fontName=self.font_name,
                    fontSize=10,
                    textColor=colors.black,
                    leading=14,
                    backColor=colors.HexColor('#f7fafc'),
                    borderColor=colors.HexColor('#e2e8f0'),
                    borderWidth=1,
                    borderPadding=10,
                    borderRadius=5
                )
            ))
        
        # 푸터
        elements.append(Spacer(1, 8*mm))
        elements.append(HRFlowable(
            width="100%",
            thickness=1,
            color=colors.gray,
            spaceBefore=10
        ))
        elements.append(Paragraph(
            f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Daily Report",
            ParagraphStyle(
                name='Footer',
                fontName=self.font_name,
                fontSize=8,
                textColor=colors.gray,
                alignment=2  # 오른쪽 정렬
            )
        ))
        
        # PDF 빌드
        doc.build(elements)
        
        return output_path


def load_from_excel(excel_path: str) -> DailyReportData:
    """Excel 파일에서 보고서 데이터 로드
    
    예상 Excel 구조:
    - Sheet1 또는 '일일보고' 시트
    - A1: 날짜
    - A2: 당일 건수, B2: 당일 수수료
    - A3: 월누계 건수, B3: 월누계 수수료
    - A5~: 미수금 (업체명, 금액)
    - 빈 행 후: 대납금 (업체명, 금액)
    """
    import openpyxl
    
    data = DailyReportData()
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # 시트 찾기
    sheet = None
    for name in ['일일보고', '보고서', 'Sheet1', wb.sheetnames[0]]:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    
    if sheet is None:
        sheet = wb.active
    
    # 데이터 읽기 (유연한 파싱)
    rows = list(sheet.iter_rows(values_only=True))
    
    # 기본 정보 찾기
    for i, row in enumerate(rows):
        if row[0] is None:
            continue
            
        cell_str = str(row[0]).strip()
        
        # 날짜 찾기
        if '날짜' in cell_str or '일자' in cell_str:
            if len(row) > 1 and row[1]:
                data.report_date = str(row[1])
        
        # 당일 실적
        elif '당일' in cell_str or '하루' in cell_str or '금일' in cell_str:
            if '건수' in cell_str and len(row) > 1:
                data.daily_count = int(row[1] or 0)
            elif '수수료' in cell_str and len(row) > 1:
                data.daily_fee = int(row[1] or 0)
            else:
                # 건수와 수수료가 같은 행에
                if len(row) > 1:
                    data.daily_count = int(row[1] or 0)
                if len(row) > 2:
                    data.daily_fee = int(row[2] or 0)
        
        # 월 누계
        elif '월' in cell_str and ('누계' in cell_str or '합계' in cell_str):
            if len(row) > 1:
                data.monthly_count = int(row[1] or 0)
            if len(row) > 2:
                data.monthly_fee = int(row[2] or 0)
        
        # 미수금 섹션
        elif '미수금' in cell_str:
            # 다음 행부터 미수금 데이터
            for j in range(i + 1, len(rows)):
                r = rows[j]
                if r[0] is None or str(r[0]).strip() == '':
                    break
                if '대납금' in str(r[0]) or '합계' in str(r[0]):
                    break
                if len(r) > 1 and r[1]:
                    try:
                        data.add_receivable(str(r[0]), int(r[1] or 0))
                    except (ValueError, TypeError):
                        continue
        
        # 대납금 섹션
        elif '대납금' in cell_str:
            for j in range(i + 1, len(rows)):
                r = rows[j]
                if r[0] is None or str(r[0]).strip() == '':
                    break
                if '합계' in str(r[0]):
                    break
                if len(r) > 1 and r[1]:
                    try:
                        data.add_advance(str(r[0]), int(r[1] or 0))
                    except (ValueError, TypeError):
                        continue
    
    wb.close()
    return data


# 테스트용 샘플 데이터 생성
def create_sample_data() -> DailyReportData:
    """테스트용 샘플 데이터"""
    data = DailyReportData()
    data.report_date = datetime.now().strftime('%Y-%m-%d')
    data.daily_count = 15
    data.daily_fee = 750000
    data.monthly_count = 285
    data.monthly_fee = 14250000
    
    data.add_receivable('(주)한세에프엔씨', 1500000)
    data.add_receivable('삼성전자(주)', 2300000)
    data.add_receivable('(주)LG생활건강', 850000)
    
    data.add_advance('현대자동차(주)', 450000)
    data.add_advance('SK하이닉스(주)', 320000)
    
    return data


if __name__ == '__main__':
    # 테스트 실행
    generator = DailyReportGenerator()
    sample_data = create_sample_data()
    output = generator.generate_pdf(sample_data, 'test_daily_report.pdf')
    print(f"PDF 생성 완료: {output}")
