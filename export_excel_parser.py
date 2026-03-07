"""
JARVIS 수출 자동화 - 엑셀 파싱 모듈
첨부된 엑셀 파일에서 수출신고에 필요한 데이터를 추출합니다.
"""

import os
from dataclasses import dataclass
from typing import Optional, Dict, Any

try:
    import openpyxl
except ImportError:
    print("[경고] openpyxl 패키지가 없습니다. pip install openpyxl")
    openpyxl = None


@dataclass
class ExportData:
    """추출된 수출 데이터"""
    품명: str = ""
    수량: float = 0
    단가: float = 0
    총액: float = 0
    중량: float = 0
    
    # AI 선택용 추가 정보
    route: str = ""  # 경로 (ICN/BKK 등)
    is_free: bool = False  # 무상 여부
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "품명": self.품명,
            "수량": self.수량,
            "단가": self.단가,
            "총액": self.총액,
            "중량": self.중량,
            "route": self.route,
            "is_free": self.is_free
        }
    
    def __str__(self) -> str:
        trade_type = "무상" if self.is_free else "유상"
        return (f"품명: {self.품명}\n"
                f"수량: {self.수량}\n"
                f"단가: {self.단가}\n"
                f"총액: {self.총액}\n"
                f"중량: {self.중량}\n"
                f"경로: {self.route}\n"
                f"거래유형: {trade_type}")


class ExportExcelParser:
    """
    수출 요청 엑셀 파일 파서
    
    엑셀 구조 (실제):
    - 시트1 (INVOICE 1): D24=품명, E24=수량, F24=단가, G24=총액
    - 시트2 (Packing List): Row 24에서 중량 값
    """
    
    # 셀 위치 상수
    INVOICE_SHEET_INDEX = 0  # 첫 번째 시트
    PACKING_SHEET_INDEX = 1  # 두 번째 시트
    
    # INVOICE 셀 위치 (Row 24)
    CELL_ITEM_NAME = "B24"   # 품명 (Description of goods)
    CELL_QUANTITY = "E24"    # 수량 (In Box)
    CELL_UNIT_PRICE = "F24"  # 단가 (Price/Box)
    CELL_TOTAL = "G24"       # 총액 (Total Amount)
    
    # Packing List 셀 위치 - Row 24에서 중량 찾기
    WEIGHT_ROW = 24          # 중량이 있는 행
    
    def __init__(self, file_path: str):
        """
        Args:
            file_path: 엑셀 파일 경로
        """
        self.file_path = file_path
        self.workbook = None
        
    def _open_workbook(self) -> bool:
        """워크북 열기"""
        if openpyxl is None:
            print("[오류] openpyxl이 설치되지 않았습니다.")
            return False
            
        if not os.path.exists(self.file_path):
            print(f"[오류] 파일을 찾을 수 없습니다: {self.file_path}")
            return False
            
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            return True
        except Exception as e:
            print(f"[오류] 엑셀 파일 열기 실패: {e}")
            return False
    
    def _get_cell_value(self, sheet_index: int, cell_address: str) -> Any:
        """특정 셀 값 가져오기"""
        try:
            if sheet_index >= len(self.workbook.sheetnames):
                print(f"[경고] 시트 인덱스 {sheet_index}가 범위를 벗어남")
                return None
                
            sheet = self.workbook.worksheets[sheet_index]
            value = sheet[cell_address].value
            return value
        except Exception as e:
            print(f"[오류] 셀 읽기 실패 (시트{sheet_index}, {cell_address}): {e}")
            return None
    
    def _parse_number(self, value: Any) -> float:
        """값을 숫자로 변환"""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        try:
            # 문자열에서 숫자 추출 (콤마, 공백 제거)
            cleaned = str(value).replace(",", "").replace(" ", "").strip()
            return float(cleaned) if cleaned else 0.0
        except ValueError:
            return 0.0
    
    def parse_invoice_sheet(self) -> Dict[str, Any]:
        """
        시트1 (COMMERCIAL INVOICE) 파싱
        
        Returns:
            {"품명": str, "수량": float, "단가": float, "총액": float}
        """
        result = {
            "품명": "",
            "수량": 0.0,
            "단가": 0.0,
            "총액": 0.0
        }
        
        # 품명
        품명 = self._get_cell_value(self.INVOICE_SHEET_INDEX, self.CELL_ITEM_NAME)
        result["품명"] = str(품명) if 품명 else ""
        
        # 수량
        수량 = self._get_cell_value(self.INVOICE_SHEET_INDEX, self.CELL_QUANTITY)
        result["수량"] = self._parse_number(수량)
        
        # 단가
        단가 = self._get_cell_value(self.INVOICE_SHEET_INDEX, self.CELL_UNIT_PRICE)
        result["단가"] = self._parse_number(단가)
        
        # 총액
        총액 = self._get_cell_value(self.INVOICE_SHEET_INDEX, self.CELL_TOTAL)
        result["총액"] = self._parse_number(총액)
        
        return result
    
    def parse_packing_sheet(self) -> Dict[str, Any]:
        """
        시트2 (Packing List) 파싱
        Row 24에서 중량 값 찾기
        
        Returns:
            {"중량": float}
        """
        result = {"중량": 0.0}
        
        try:
            sheet = self.workbook.worksheets[self.PACKING_SHEET_INDEX]
            # Row 24에서 숫자 값들 중 마지막(또는 큰 값)을 중량으로 간주
            row = sheet[self.WEIGHT_ROW]
            for cell in reversed(row):
                value = cell.value
                if value is not None and isinstance(value, (int, float)):
                    result["중량"] = float(value)
                    break
        except Exception as e:
            print(f"[오류] Packing List 파싱 실패: {e}")
        
        return result
    
    def parse_all(self) -> Optional[ExportData]:
        """
        전체 엑셀 파싱
        
        Returns:
            ExportData 객체 또는 실패 시 None
        """
        if not self._open_workbook():
            return None
            
        try:
            # 시트1 파싱
            invoice_data = self.parse_invoice_sheet()
            
            # 시트2 파싱
            packing_data = self.parse_packing_sheet()
            
            # 결과 조합
            return ExportData(
                품명=invoice_data["품명"],
                수량=invoice_data["수량"],
                단가=invoice_data["단가"],
                총액=invoice_data["총액"],
                중량=packing_data["중량"]
            )
            
        except Exception as e:
            print(f"[오류] 파싱 실패: {e}")
            return None
        finally:
            if self.workbook:
                self.workbook.close()


# 테스트용
if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("사용법: python export_excel_parser.py <엑셀파일경로>")
        print("\n테스트용 더미 데이터 생성 중...")
        
        # 테스트용 엑셀 파일 생성
        if openpyxl:
            wb = openpyxl.Workbook()
            
            # 시트1: COMMERCIAL INVOICE
            ws1 = wb.active
            ws1.title = "COMMERCIAL INVOICE"
            ws1["A1"] = "COMMERCIAL INVOICE"
            ws1["B25"] = "삼성전자 반도체 부품"
            ws1["E25"] = 100
            ws1["F25"] = 50.5
            ws1["G25"] = 5050
            
            # 시트2: Packing List
            ws2 = wb.create_sheet("Packing List")
            ws2["A1"] = "Packing List"
            ws2["L25"] = 25.5
            
            test_file = "test_export.xlsx"
            wb.save(test_file)
            wb.close()
            
            print(f"테스트 파일 생성됨: {test_file}")
            print("\n파싱 테스트:")
            
            parser = ExportExcelParser(test_file)
            result = parser.parse_all()
            
            if result:
                print("=== 파싱 결과 ===")
                print(result)
            else:
                print("파싱 실패")
                
            # 테스트 파일 삭제
            os.remove(test_file)
            print(f"\n테스트 파일 삭제됨: {test_file}")
    else:
        file_path = sys.argv[1]
        parser = ExportExcelParser(file_path)
        result = parser.parse_all()
        
        if result:
            print("=== 파싱 결과 ===")
            print(result)
        else:
            print("파싱 실패")
